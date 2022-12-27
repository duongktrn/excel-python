import pandas as pd
import openpyxl as op 
import openpyxl
# 2 thư viện ở import ảnh vào excel
from openpyxl.drawing.image import Image
import os
#đọc số tiền thành chữ tiền 
from num2words import num2words
# cách để git lên github
#1. tải git về máy tính ,sau đó gõ lệnh git init
#2 .git remote add origin https://github.com/duongktrn/excel-python.git
#3.   git config --global user.email "you@example.com" 
#4. git config --global user.name "Your Name"
# git add . (để add tất cả file)
# git commit - m "tên file" để commit file
# git push origin master 


# ############## 1. pandas

# #Đọc file excel bằng datafarm
# df = pd.read_excel('pd.xlsx')

# #đọc dữ liệu 1 cột (lấy cả giá trị rỗng ,và index lớn nhất của 1 cột) :
# column_A = df['name']
# column_B = df['address']
# column_C = df['phone']

# #đọc dữ liệu 1 cột bỏ qua giá trị none của hàng:
# column_D = df['address'].dropna()

# # ghi dữ liệu 1 cột vào 1 mảng :
# arr_A = [column for column in column_A]
# arr_B = [column for column in column_B]
# arr_C = [column for column in column_C]
# arr_D = [column for column in column_D]
# print(arr_A)
# print(arr_B)
# print(arr_C)
# print(arr_D)

# #Tìm vị trí index của 1 ô trong 1 cột:
# index=df[df['name']=='nga'].index[0]
# print(index)

# # Tìm vị trí index cuối cùng :
# print(column_C.index.stop)
# # cách 2
# index = df.index.stop
# print(index)

# # Thêm giá trị vào ô 
# df.at[7,'phone'] =999
# df.to_excel('pd.xlsx',index=False)

# # Lấy giá trị của 1 ô:
# value_A = df.at[2,'name']
# print(value_A)
# # dùng phương thức loc
# #khác nhau giữ loc và at là at xử lý nhanh hơn ,nhưng loc có thể lấy 1 lúc nhiều chỉ mục index
# value_B = df.loc[2,'address']
# print(value_B)
# value_C = df.loc[[1,2,3],'address']
# print(value_C)

# #ghi dữ liệu 1 mảng vào 1 cột ,chú ý : index trong mảng phải bằng index cột ,nếu thiếu thì phải thêm giá trị rỗng vào mảng để bằng nhau
# #cách này nhanh hơn dùng phương thức at()
# df['test'] = arr_A
# print(df['test'])

# # head(): trả về các hàng đầu tiên của dataframe.
# # tail(): trả về các hàng cuối cùng của dataframe.
# # shape: trả vế số hàng và cột của dataframe.
# # columns: trả về danh sách tên các cột trong dataframe.
# # info(): trả về thông tin chi tiết về dataframe, bao gồm số lượng hàng, số lượng cột, kiểu dữ liệu của từng cột, v.v.

# # sort_values(): sắp xếp dataframe theo một hoặc nhiều cột nhất định.
# # groupby(): nhóm dữ liệu trong dataframe theo một hoặc nhiều cột nhất định.
# # pivot_table(): tạo bảng chuyển đổi (pivot table) từ dataframe.
# # reset_index(): thiết lập lại chỉ mục của dataframe.
# # drop(): xóa một hoặc nhiều hàng hoặc cột khỏi dataframe.
# # rename(): đổi tên một hoặc nhiều cột trong dataframe.
# # fillna(): điền giá trị mặc định vào các ô bị thiếu dữ liệu trong dataframe.
# df['phone']=df['phone'].fillna(0)
# df = df.fillna(0)
# print(df['phone'])
# print(df)


# ########################### 2  openpyxl
wb_excel = openpyxl.load_workbook('op.xlsx')
sheet = wb_excel.active

# số hàng lớn nhất 
max_row = sheet.max_row
print(max_row)


## lấy só hàng của 1 cột không bao gồm None
# Cột muốn lấy số hàng
column = 'B'
# Biến lưu số hàng trong cột
num_rows = 0
# Duyệt qua tất cả các ô trong cột
for row in range(1, sheet.max_row + 1):
  cell = sheet[column + str(row)]
  if cell.value is not None:
    num_rows += 1
print(num_rows) 

#Lấy các giá trị của 1 cột bao gồm cả None
column = sheet['A']
value_a = [cell.value for cell in column]
print(value_a)

#Lấy giá trị của 1 ô bất kỳ :
value_b = sheet['b6'].value
print(value_b)
# cách 2:
value_name = sheet.cell(row=6, column=2)
print(value_name.value)

# Ghi giá trị vào 1 ô :
cell_name = sheet.cell(row=3, column=3)
cell_name.value = 'hello'
wb_excel.save('op.xlsx')
print(cell_name.value)

#Đọc số tiền thành chữ:
number = 100000
words = num2words(number,lang='vi')
print(words)

#Duyệt các ô trong 1 khoảng ô nhất định bằng inter_row(),chậm hơn iter_col hoặc value:
for row in sheet.iter_rows(min_row=1,max_row=5,min_col=1,max_col=2):
    arr = [cell.value for cell in row]
    print(arr)


#thêm ảnh vào file excel
pathImage = 'anh.jpeg'
#dùng os để kiểm tra ảnh có tồn tại hay không
if os.path.exists(pathImage):
    draw = Image(pathImage)
    draw.anchor='D1'
    sheet.add_image(draw)
    wb_excel.save('op.xlsx')
else:
    print('khong ton tai path')